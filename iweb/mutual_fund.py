import math
import openpyxl
from re import findall
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.webdriver import WebDriver
from utils import find_element_or_none


def iweb_mutual_fund(DRIVER: WebDriver, out_xlsx: str) -> None:
    url = 'https://www.markets.iweb-sharedealing.co.uk/funds-centre/'
    WAIT = WebDriverWait(DRIVER, 10)
    DRIVER.get(url)

    total_elm = find_element_or_none(WAIT, '//*[@id="fullFundRangeTotal"]')
    if not total_elm:
        print("total_elm not found")
        return
    print(total_elm)
    total = total_elm.text.replace(',', '')
    total_funds = int(findall(r'[0-9]+', total)[0])
    total_pages = math.ceil(total_funds / 100)

    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb['MF']
    sheet_iter = 2

    print(f'[iWeb] Total MF = {total_funds}')
    for page in range(1, total_pages + 1):
        data = []
        current_page = f'https://www.markets.iweb-sharedealing.co.uk/modules/funds/full-fund-range-result/?orderField=UnitNameLong&orderType=asc&offset={page}&limit=100'
        DRIVER.get(current_page)
        print(f'[#] MF Page {page}/{total_pages} [#]')
        WAIT.until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/table/tbody/tr')))
        tr = DRIVER.find_elements(By.XPATH, '/html/body/table/tbody/tr')
        even_tr = tr[::2]

        for row in even_tr:
            name = row.find_element(By.XPATH, './td[1]/p/a').text.strip()
            url = row.find_element(
                By.XPATH, './td[1]/p/a').get_attribute('href')

            isin = ""
            if url:
                isin_match = findall(r"[A-Z0-9]{12}", url)
                if len(isin_match) > 0:
                    isin = isin_match[0]
            data.append({'name': name, 'url': url, 'isin': isin})

        for output in data:
            ws.cell(sheet_iter, 1).value = output['name']
            ws.cell(sheet_iter, 2).value = output["isin"]
            fund_url = output["url"]
            c = ws.cell(sheet_iter, 3, fund_url)
            c.hyperlink = fund_url
            c.style = "Hyperlink"
            sheet_iter += 1
        wb.save(out_xlsx)

    wb.save(out_xlsx)
    wb.close()
